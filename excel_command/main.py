import msoffcrypto
import openpyxl
import json
import io
from datetime import datetime
import requests
import time

def send_data_to_esp32(json_data):
    
    # ESP32 server URL
    esp32_url = "http://172.17.8.33/receive_data"

    try:
        response = requests.post(esp32_url, json=json_data)
        
        if response.status_code == 200:
            print("JSON data sent successfully to ESP32.")
        else:
            print(f"Failed to send JSON data to ESP32. Status code: {response.status_code}")
    except Exception as e:
        print(f"An error occurred while sending JSON data to ESP32: {str(e)}")

# Main function to continuously read Excel and send data to ESP32 every 10 seconds
def main():
    # Decrypt the workbook
    workbook_path = r'Suivi_des_Moules.xlsx'
    password = 'Gree123'

    while True:
        try:
            decrypted_stream = io.BytesIO()

            with open(workbook_path, 'rb') as f:
                file = msoffcrypto.OfficeFile(f)
                file.load_key(password=password)
                file.decrypt(decrypted_stream)

            decrypted_stream.seek(0)  
            workbook = openpyxl.load_workbook(decrypted_stream, read_only=True)

            target_sheet = "Injection plastique "

            if target_sheet in workbook.sheetnames:
                worksheet = workbook[target_sheet]
                
                # Initialize sheet_data dictionary
                sheet_data = {}
                
                for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=6):
                    identifier = row[0].value  # Column B
                    value = row[4].value       # Column F
                    if identifier is not None and value is not None:
                        if isinstance(identifier, datetime):
                            identifier = identifier.strftime('%Y-%m-%d %H:%M:%S')
                        else:
                            identifier = str(identifier)  

                        sheet_data[identifier] = {"Status": value}
                
                json_data = {target_sheet: sheet_data}

                send_data_to_esp32(json_data)
            else:
                print(f"Sheet '{target_sheet}' not found in the workbook.")
            
            decrypted_stream.close()

            time.sleep(10)
        except Exception as e:
            print(f"An error occurred: {str(e)}")
            time.sleep(10)

if __name__ == "__main__":
    main()
